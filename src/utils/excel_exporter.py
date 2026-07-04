import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generar_reporte_excel_como_bytes(productos, historial):
    buffer = io.BytesIO()
    exito = generar_reporte_excel(buffer, productos, historial)
    if exito:
        buffer.seek(0)
    return buffer if exito else None

def generar_reporte_completo_excel(ruta: str) -> bool:
    try:
        from src.database.conexion import obtener_conexion
        conexion = obtener_conexion()
        cursor = conexion.cursor()
        cursor.execute("SELECT nombre, cantidad, precio_costo, detalles, stock_minimo FROM productos")
        productos = []
        for row in cursor.fetchall():
            productos.append({
                "nombre": row[0], "cantidad": row[1],
                "precio_costo": row[2], "detalles": row[3] or "",
                "stock_minimo": row[4]
            })
        cursor.execute("SELECT fecha_hora, tipo_evento, descripcion FROM historial_movimientos ORDER BY id DESC")
        historial = []
        for row in cursor.fetchall():
            historial.append({"fecha": row[0], "tipo": row[1], "descripcion": row[2]})
        conexion.close()
        with open(ruta, "wb") as f:
            buffer = io.BytesIO()
            ok = generar_reporte_excel(buffer, productos, historial)
            if not ok:
                return False
            f.write(buffer.getvalue())
        return True
    except Exception as e:
        print(f"Error generando reporte completo: {e}")
        return False

def generar_reporte_excel(archivo_salida, productos, historial) -> bool:
    try:
        libro = openpyxl.Workbook()

        fuente_cabecera = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fuente_total = Font(name="Calibri", size=11, bold=True, color="000000")
        relleno_cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        relleno_total = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        alineacion_centro = Alignment(horizontal="center", vertical="center")

        ws_kardex = libro.active
        ws_kardex.title = "Libro Contable"
        cabeceras = ["Fecha y Hora", "Detalle / Operación", "Cantidad Movida",
                     "Precio Unitario ($)", "Entradas (Costo Total) ($)",
                     "Salidas (Venta Total) ($)", "Saldo Neto ($)"]
        ws_kardex.append(cabeceras)

        fila_actual = 2
        for h in historial:
            cantidad_piezas = 0
            precio_unitario = 0.0
            entrada_total = 0.0
            salida_total = 0.0

            try:
                desc = h.get("descripcion", "")
                tipo = h.get("tipo", "")
                if tipo in ("CREACION", "ACTUALIZACION") and "cantidad:" in desc.lower():
                    partes = desc.split("|")
                    for parte in partes:
                        if "cantidad:" in parte.lower():
                            cantidad_piezas = int(parte.split(":")[-1].replace("uds", "").strip())
                        if "costo unitario:" in parte.lower():
                            precio_unitario = float(parte.split("$")[-1].strip())
                    entrada_total = cantidad_piezas * precio_unitario
                elif tipo == "VENTA":
                    partes = desc.split("|")
                    for parte in partes:
                        if "cantidad:" in parte.lower():
                            cantidad_piezas = int(parte.split(":")[-1].replace("uds", "").strip())
                        if "precio venta:" in parte.lower():
                            precio_unitario = float(parte.split("$")[-1].strip())
                    salida_total = cantidad_piezas * precio_unitario
            except (ValueError, IndexError, AttributeError):
                pass

            saldo_neto = salida_total - entrada_total
            ws_kardex.append([
                h.get("fecha", ""), desc, cantidad_piezas,
                precio_unitario, entrada_total, salida_total, saldo_neto
            ])
            fila_actual += 1

        ws_kardex.append([
            "TOTALES", "Sumatoria global analítica", "-", "-",
            f"=SUM(E2:E{fila_actual-1})",
            f"=SUM(F2:F{fila_actual-1})",
            f"=SUM(G2:G{fila_actual-1})"
        ])

        for col_idx in range(1, 8):
            celda_tot = ws_kardex.cell(row=fila_actual, column=col_idx)
            celda_tot.font = fuente_total
            celda_tot.fill = relleno_total

        ws_inventario = libro.create_sheet(title="Inventario Disponible")
        ws_inventario.append(["ID", "Nombre", "Cantidad en Stock",
                              "Costo Adquisición ($)", "Detalles", "Stock Mínimo"])

        for i, p in enumerate(productos, 1):
            ws_inventario.append([
                i, p.get("nombre", ""), p.get("cantidad", 0),
                p.get("precio_costo", 0.0), p.get("detalles", ""),
                p.get("stock_minimo", 5)
            ])

        for hoja in libro.worksheets:
            for col_idx in range(1, hoja.max_column + 1):
                celda = hoja.cell(row=1, column=col_idx)
                celda.font = fuente_cabecera
                celda.fill = relleno_cabecera
                celda.alignment = alineacion_centro

            for col in hoja.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                hoja.column_dimensions[col_letter].width = max(max_len + 4, 16)

        libro.save(archivo_salida)
        return True

    except Exception as e:
        print(f"Error generando Excel: {e}")
        return False
