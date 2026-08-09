import io

import openpyxl

from src.utils.excel_exporter import generar_reporte_excel


class TestGenerarReporteExcel:
    def test_con_datos_devuelve_true(self):
        buffer = io.BytesIO()
        productos = [{"nombre": "Laptop", "cantidad": 10, "precio_costo": 500.0, "detalles": "", "stock_minimo": 2}]
        historial = [
            {
                "fecha": "2024-01-01 10:00",
                "tipo": "CREACION",
                "descripcion": "Ingreso: Laptop | Cantidad: 5 uds | Costo Unitario: $500",
            }
        ]
        assert generar_reporte_excel(buffer, productos, historial) is True

    def test_sin_productos_devuelve_true(self):
        buffer = io.BytesIO()
        assert generar_reporte_excel(buffer, [], []) is True

    def test_archivo_tiene_hojas_correctas(self):
        buffer = io.BytesIO()
        productos = [{"nombre": "Mouse", "cantidad": 20, "precio_costo": 15.0, "detalles": "", "stock_minimo": 5}]
        historial = []
        generar_reporte_excel(buffer, productos, historial)
        wb = openpyxl.load_workbook(buffer)
        assert "Libro Contable" in wb.sheetnames
        assert "Inventario Disponible" in wb.sheetnames

    def test_historial_con_varios_tipos(self):
        buffer = io.BytesIO()
        productos = [{"nombre": "Teclado", "cantidad": 8, "precio_costo": 30.0, "detalles": "", "stock_minimo": 3}]
        historial = [
            {
                "fecha": "2024-01-01",
                "tipo": "CREACION",
                "descripcion": "Ingreso: Teclado | Cantidad: 10 uds | Costo Unitario: $30",
            },
            {
                "fecha": "2024-01-02",
                "tipo": "VENTA",
                "descripcion": "Salida de stock | Cantidad: 2 uds | Precio Venta: $45 | Total: $90.00",
            },
            {"fecha": "2024-01-03", "tipo": "BAJO STOCK", "descripcion": "Stock bajo para: Teclado"},
        ]
        assert generar_reporte_excel(buffer, productos, historial) is True

    def test_multiples_productos(self):
        buffer = io.BytesIO()
        productos = [
            {
                "nombre": f"Producto {i}",
                "cantidad": i * 10,
                "precio_costo": float(i * 100),
                "detalles": "",
                "stock_minimo": 5,
            }
            for i in range(1, 21)
        ]
        historial = []
        assert generar_reporte_excel(buffer, productos, historial) is True


class TestContenidoLibroContable:
    def _generar_libro(self, productos, historial):
        buffer = io.BytesIO()
        generar_reporte_excel(buffer, productos, historial)
        buffer.seek(0)
        return openpyxl.load_workbook(buffer)

    def test_inventario_contiene_datos_correctos(self):
        productos = [{"nombre": "Laptop", "cantidad": 10, "precio_costo": 500.0, "detalles": "i7", "stock_minimo": 2}]
        wb = self._generar_libro(productos, [])
        ws = wb["Inventario Disponible"]
        assert ws.cell(row=2, column=2).value == "Laptop"
        assert ws.cell(row=2, column=3).value == 10
        assert ws.cell(row=2, column=4).value == 500.0

    def test_fila_totales_incluye_formulas(self):
        historial = [
            {
                "fecha": "2024-01-01",
                "tipo": "CREACION",
                "descripcion": "Ingreso: Laptop | Cantidad: 5 uds | Costo Unitario: $500",
            },
            {
                "fecha": "2024-01-02",
                "tipo": "VENTA",
                "descripcion": "Salida de stock | Cantidad: 2 uds | Precio Venta: $450",
            },
        ]
        wb = self._generar_libro([], historial)
        ws = wb["Libro Contable"]
        filas = ws.max_row
        assert ws.cell(row=filas, column=1).value == "TOTALES"
        assert str(ws.cell(row=filas, column=5).value).startswith("=SUM(E2:E")
        assert str(ws.cell(row=filas, column=6).value).startswith("=SUM(F2:F")

    def test_venta_parsea_montos_en_kardex(self):
        historial = [
            {
                "fecha": "2024-01-02",
                "tipo": "VENTA",
                "descripcion": "Salida de stock | Cantidad: 3 uds | Precio Venta: $100",
            },
        ]
        wb = self._generar_libro([], historial)
        ws = wb["Libro Contable"]
        fila = 2
        assert ws.cell(row=fila, column=3).value == 3
        assert ws.cell(row=fila, column=4).value == 100.0
        assert ws.cell(row=fila, column=6).value == 300.0

    def test_creacion_parsea_entradas_en_kardex(self):
        historial = [
            {
                "fecha": "2024-01-01",
                "tipo": "CREACION",
                "descripcion": "Ingreso: Laptop | Cantidad: 4 uds | Costo Unitario: $50",
            },
        ]
        wb = self._generar_libro([], historial)
        ws = wb["Libro Contable"]
        fila = 2
        assert ws.cell(row=fila, column=5).value == 200.0
        assert ws.cell(row=fila, column=7).value == -200.0
